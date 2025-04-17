from fastapi import APIRouter, File, UploadFile
from pydantic import BaseModel
from openpyxl import load_workbook
from fastapi.responses import FileResponse
import time
import os


app = APIRouter()

class ReportData(BaseModel):
    full_name: str
    department: str
    id: int
    rang: str
    ticket_date: str
    ticket_type: str
    ticket_price: float
    id_ticket: str
    hotel_date: str
    hotel_name: str
    hotel_price: float
    daily_allowance: float

TEMPLATE_PATH = "./template.xlsx"
OUTPUT_PATH = "./advance_report.xlsx"
DATE = time.localtime()


@app.post('/create_report')
async def create_report(data: ReportData):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    ws["J13"] = data.full_name
    ws["I10"] = data.department
    ws["AU13"] = data.id
    ws["G15"] = data.rang
    ws["C26"] = data.ticket_date
    ws["G26"] = data.id_ticket
    ws["M26"] = data.ticket_type
    ws["Y26"] = data.ticket_price
    ws["C25"] = data.hotel_date
    #ws["C8"] = data.hotel_name
    ws["Y25"] = data.hotel_price
    ws["Y24"] = data.daily_allowance
    #ws["D10"] = data.ticket_price + data.hotel_price + data.daily_allowance
    ws["W7"] = f"{DATE.tm_mday}.{DATE.tm_mon}.{DATE.tm_year}"

    wb.save(OUTPUT_PATH)
    wb.close()

    return FileResponse(OUTPUT_PATH, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename="Advance_Report.xlsx")

