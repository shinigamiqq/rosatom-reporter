from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook
from fastapi.responses import FileResponse
import time
import os


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class TicketItem(BaseModel):
    filename: str
    Date: str
    Price: float
    Ticket: str
    TicketType: str

class HotelItem(BaseModel):
    filename: str
    Date: str
    Price: float

class ReportData(BaseModel):
    full_name: str
    department: str
    id: int
    rang: str
    daily_allowance: float
    tickets: list[TicketItem]
    hotels: list[HotelItem]

TEMPLATE_PATH = "./template.xlsx"
OUTPUT_PATH = "./advance_report.xlsx"
DATE = time.localtime()

@app.post('/create_report')
async def create_report(data: ReportData):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    ws["J13"] = data.full_name
    ws["I10"] = data.department
    ws["AU13"] = data.id
    ws["G15"] = data.rang
    ws["Y24"] = data.daily_allowance
    ws["W7"] = f"{DATE.tm_mday}.{DATE.tm_mon}.{DATE.tm_year}"

    start_row = 25

    for i, ticket in enumerate(data.tickets):
        row = start_row + i
        ws[f"C{row}"] = ticket.Date
        ws[f"G{row}"] = ticket.Ticket
        ws[f"M{row}"] = ticket.TicketType
        ws[f"Y{row}"] = ticket.Price

    offset = len(data.tickets)
    for i, hotel in enumerate(data.hotels):
        row = start_row + offset + i
        ws[f"C{row}"] = hotel.Date
        ws[f"M{row}"] = "Проживание"
        ws[f"Y{row}"] = hotel.Price

    wb.save(OUTPUT_PATH)
    wb.close()

    return FileResponse(OUTPUT_PATH, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename="Advance_Report.xlsx")

